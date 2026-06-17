#!/usr/bin/env python3
"""Infer standardized genomics source metadata from arbitrary local raw files."""

from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

import yaml

from genomics_contract import items_for_header, load_contract, promoted_entities_for_header, split_shared_and_specific
from normalize_source import GENE_COLUMNS, PROTEIN_COLUMNS, TRANSCRIPT_COLUMNS, derive_contrast_summary_links


CONTRACT = load_contract("functional_genomics")


def _expression_label_from_column(column: str, prefix: str) -> str:
    core = column[len(prefix):] if column.startswith(prefix) else column
    parts = [part for part in core.split("_") if part]
    if prefix == "dge_" and len(parts) >= 2:
        return f"{parts[0]} vs {' '.join(parts[1:])}"
    return " ".join(parts) if parts else column


def _infer_expression_entities(header: list[str]) -> dict[str, Any]:
    expression_defaults = CONTRACT.get("expression_entities", {}) or {}
    summaries_defaults = expression_defaults.get("summaries_defaults", {}) or {}
    contrasts_defaults = expression_defaults.get("contrasts_defaults", {}) or {}
    expression_fields = {
        _expression_label_from_column(column, "avg_"): column
        for column in header
        if column.startswith("avg_")
    }
    summaries = [
        {
            "label": _expression_label_from_column(column, "avg_"),
            "column": column,
            "attach_to": "transcript",
            "measure_type": "summary",
            "order_index": index,
            "entity_type": summaries_defaults.get("entity_type", ""),
            "relation_type": summaries_defaults.get("relation_type", ""),
            "id_template": summaries_defaults.get("id_template", ""),
            "value_key": summaries_defaults.get("value_key", "value"),
            "parent_tag": summaries_defaults.get("parent_tag", ""),
        }
        for index, column in enumerate(column for column in header if column.startswith("avg_"))
    ]
    contrasts = [
        {
            "label": _expression_label_from_column(column, "dge_"),
            "column": column,
            "attach_to": "transcript",
            "measure_type": "contrast",
            "order_index": index,
            "entity_type": contrasts_defaults.get("entity_type", ""),
            "relation_type": contrasts_defaults.get("relation_type", ""),
            "id_template": contrasts_defaults.get("id_template", ""),
            "value_key": contrasts_defaults.get("value_key", "value"),
            "parent_tag": contrasts_defaults.get("parent_tag", ""),
            **derive_contrast_summary_links(
                expression_fields,
                _expression_label_from_column(column, "dge_"),
                column,
            ),
        }
        for index, column in enumerate(column for column in header if column.startswith("dge_"))
    ]
    return {
        "summaries": split_shared_and_specific(summaries, contract_items=[]),
        "contrasts": split_shared_and_specific(contrasts, contract_items=[]),
    }


def _infer_orthogroup_promoted_entities(header: list[str]) -> dict[str, Any]:
    header_set = set(header)
    if "orthogroup" not in header_set:
        return {}
    metadata_columns = [
        column for column in [
            "glycines_gene_count",
            "schachtii_gene_count",
            "schachtii_genes",
            "schachtii_hits",
        ]
        if column in header_set
    ]
    return {
        "orthogroup": {
            "source_column": "orthogroup",
            "entity_type": "orthogroup",
            "id_template": "orthogroup:{value}",
            "name_template": "{value}",
            "attach_from": "gene",
            "relationship_type": "BELONGS_TO_ORTHOGROUP",
            "metadata_columns": metadata_columns,
        }
    }


def _infer_comparative_entities(header: list[str]) -> dict[str, Any]:
    header_set = set(header)
    entities: dict[str, Any] = {}
    if "schachtii_genes" in header_set:
        entities["homolog_family_member"] = {
            "source_column": "schachtii_genes",
            "entity_type": "bcn_gene",
            "id_template": "bcn_gene:heterodera-schachtii:{value}",
            "name_template": "{value}",
            "attach_from": "orthogroup",
            "relationship_type": "HAS_BCN_MEMBER",
            "target_organism": "Heterodera schachtii",
            "scope_tag_id": "homology-scope-cyst-nematode",
            "parser": "term_list",
        }
    if "schachtii_hits" in header_set:
        entities["bcn_hit"] = {
            "source_column": "schachtii_hits",
            "entity_type": "comparative_hit",
            "id_template": "comparative_hit:cyst_nematode:{value}",
            "name_template": "{value}",
            "attach_from": "protein",
            "relationship_type": "HAS_BCN_HIT",
            "target_organism": "Heterodera schachtii",
            "reuse_target_organism_identity": True,
            "scope_tag_id": "homology-scope-cyst-nematode",
            "parser": "term_list",
        }
    if "celegans_hits" in header_set:
        entities["nematode_hit"] = {
            "source_column": "celegans_hits",
            "entity_type": "comparative_hit",
            "id_template": "comparative_hit:nematode:{value}",
            "name_template": "{value}",
            "attach_from": "protein",
            "relationship_type": "HAS_NEMATODE_HIT",
            "scope_tag_id": "homology-scope-nematode",
            "parser": "term_list",
        }
    if "sp_best_hit" in header_set:
        entities["sp_best_hit"] = {
            "source_column": "sp_best_hit",
            "entity_type": "comparative_hit",
            "id_template": "comparative_hit:broad_parasitism:{value}",
            "name_template": "{value}",
            "attach_from": "protein",
            "relationship_type": "HAS_BROAD_HOMOLOGY_HIT",
            "scope_tag_id": "homology-scope-broad-parasitism",
            "parser": "term_list",
        }
    if "nr_best_hit" in header_set:
        entities["nr_best_hit"] = {
            "source_column": "nr_best_hit",
            "entity_type": "comparative_hit",
            "id_template": "comparative_hit:broad_parasitism:{value}",
            "name_template": "{value}",
            "attach_from": "protein",
            "relationship_type": "HAS_BROAD_HOMOLOGY_HIT",
            "scope_tag_id": "homology-scope-broad-parasitism",
            "parser": "term_list",
        }
    return entities


GENE_HINTS = ("gene", "genome", "locus", "chrom", "copy_number", "copynumber", "nested")
TRANSCRIPT_HINTS = ("transcript", "mrna", "expression", "avg_", "dge_", "cluster", "bin", "tpm", "count")
PROTEIN_HINTS = (
    "protein",
    "peptide",
    "signal",
    "localiz",
    "interpro",
    "pfam",
    "smart",
    "panther",
    "funfam",
    "go_",
    "deepgo",
    "domain",
    "disorder",
    "mol_weight",
    "isoel",
    "charge",
    "acidic",
    "basic",
    "polar",
    "aromatic",
    "alanine",
    "glycine",
    "effector",
    "orthogroup",
    "hit",
    "hgt",
)


def _slug(text: str) -> str:
    lowered = text.strip().lower()
    lowered = re.sub(r"[^a-z0-9]+", "-", lowered)
    lowered = re.sub(r"-+", "-", lowered).strip("-")
    return lowered or "genomics-dataset"


def _detect_delimiter(path: Path) -> str:
    if path.suffix.lower() == ".tsv":
        return "\t"
    if path.suffix.lower() == ".csv":
        sample = path.read_text(errors="replace")[:4096]
        try:
            return csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
        except csv.Error:
            return ","
    return "\t"


def _read_rows(source_file: Path) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = source_file.suffix.lower()
    if suffix in {".tsv", ".csv", ".txt"}:
        delimiter = _detect_delimiter(source_file)
        with source_file.open(newline="") as handle:
            rows = list(csv.DictReader(handle, delimiter=delimiter))
        return (list(rows[0].keys()) if rows else []), rows
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(source_file, read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return [], []
        header = [str(cell).strip() if cell is not None else "" for cell in values[0]]
        rows: list[dict[str, Any]] = []
        for row_values in values[1:]:
            row = {}
            for idx, col in enumerate(header):
                if not col:
                    continue
                row[col] = row_values[idx] if idx < len(row_values) else None
            rows.append(row)
        return header, rows
    raise ValueError(f"Unsupported source file type: {source_file.suffix}")


def _parse_sidecar(path: Path) -> dict[str, str]:
    suffix = path.suffix.lower()
    text = path.read_text(errors="replace")
    info: dict[str, str] = {}
    if suffix in {".yaml", ".yml"}:
        data = yaml.safe_load(text) or {}
        if isinstance(data, dict):
            for key in ("dataset_name", "name", "title", "organism", "species", "description"):
                if data.get(key):
                    info[key] = str(data[key])
        return info
    if suffix == ".json":
        data = json.loads(text)
        if isinstance(data, dict):
            for key in ("dataset_name", "name", "title", "organism", "species", "description"):
                if data.get(key):
                    info[key] = str(data[key])
        return info
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("#") and "title" not in info:
            info["title"] = stripped.lstrip("#").strip()
            continue
        match = re.match(r"(?i)^(dataset[_ ]name|name|title|organism|species|description)\s*:\s*(.+)$", stripped)
        if match:
            info[match.group(1).lower().replace(" ", "_")] = match.group(2).strip()
    return info


def _pick_context(note_paths: list[Path], source_file: Path) -> dict[str, str]:
    info: dict[str, str] = {}
    for note_path in note_paths:
        if note_path.exists():
            info.update(_parse_sidecar(note_path))
    if "dataset_name" not in info:
        info["dataset_name"] = info.get("title") or source_file.stem.replace("_", " ").replace("-", " ").title()
    if "organism" not in info and "species" in info:
        info["organism"] = info["species"]
    return info


def _distinct_nonempty(rows: list[dict[str, Any]], column: str) -> int:
    values = {
        str(value).strip()
        for row in rows
        for value in [row.get(column)]
        if value not in (None, "")
    }
    return len(values)


def _assign_column(column: str, rows: list[dict[str, Any]]) -> str:
    lowered = column.lower()
    if column in GENE_COLUMNS:
        return "gene"
    if column in TRANSCRIPT_COLUMNS:
        return "transcript"
    if column in PROTEIN_COLUMNS:
        return "protein"
    if "sequence" in lowered:
        sample = next((str(row.get(column)).strip() for row in rows if row.get(column)), "")
        if re.fullmatch(r"[ACGTUNacgtun]+", sample[:200] or ""):
            return "transcript"
        if re.fullmatch(r"[A-Z*]+", sample[:200] or ""):
            return "protein"
    if any(hint in lowered for hint in PROTEIN_HINTS):
        return "protein"
    if any(hint in lowered for hint in TRANSCRIPT_HINTS):
        return "transcript"
    if any(hint in lowered for hint in GENE_HINTS):
        return "gene"
    return "transcript"


def _infer_primary_record(header: list[str], rows: list[dict[str, Any]]) -> tuple[str, str, str | None]:
    candidates = [
        ("transcript", "uniquename", "gene_name"),
        ("transcript", "transcript_id", "gene_id"),
        ("transcript", "feature_id", "gene_name"),
        ("gene", "gene_name", None),
        ("gene", "gene_id", None),
    ]
    row_count = len(rows)
    for entity_type, id_column, gene_column in candidates:
        if id_column in header and _distinct_nonempty(rows, id_column) >= max(1, row_count - 1):
            return entity_type, id_column, gene_column if gene_column in header else None
    return "transcript", header[0] if header else "id", "gene_name" if "gene_name" in header else None


def _infer_groups(header: list[str], rows: list[dict[str, Any]], id_column: str, gene_column: str | None) -> dict[str, list[str]]:
    groups = {"gene": [], "transcript": [], "protein": []}
    for column in header:
        if column == id_column:
            continue
        if gene_column and column == gene_column:
            continue
        groups[_assign_column(column, rows)].append(column)
    return groups


def _infer_annotation_bins(header: list[str]) -> list[dict[str, Any]]:
    return items_for_header(list(CONTRACT.get("annotation_bins", [])), header)


def _infer_tag_bins(header: list[str]) -> list[dict[str, Any]]:
    return items_for_header(list(CONTRACT.get("tag_bins", [])), header)


def _infer_boolean_tags(header: list[str]) -> list[dict[str, Any]]:
    return items_for_header(list(CONTRACT.get("boolean_tags", [])), header)


def _infer_value_presence_tags(header: list[str]) -> list[dict[str, Any]]:
    return items_for_header(list(CONTRACT.get("value_presence_tags", [])), header)


def _default_dataset_name(source_file: Path, sidecar: dict[str, str]) -> str:
    return sidecar.get("dataset_name") or source_file.stem.replace("_", " ").replace("-", " ").title()


def infer_source_package(
    *,
    source_file: Path,
    source_dir: Path | None = None,
    dataset_id: str | None = None,
    dataset_name: str | None = None,
    organism: str | None = None,
    note_paths: list[Path] | None = None,
    apply: bool = False,
) -> tuple[Path, Path, Path]:
    source_file = source_file.resolve()
    output_dir = (source_dir.resolve() if source_dir else source_file.parent.resolve())
    output_dir.mkdir(parents=True, exist_ok=True)
    notes = [path.resolve() for path in (note_paths or [])]

    header, rows = _read_rows(source_file)
    if not header:
        raise ValueError(f"No header detected in {source_file}")

    sidecar = _pick_context(notes, source_file)
    inferred_dataset_id = dataset_id or _slug(source_file.stem)
    inferred_dataset_name = dataset_name or _default_dataset_name(source_file, sidecar)
    inferred_organism = organism or sidecar.get("organism") or "Unknown organism"

    primary_entity, id_column, gene_column = _infer_primary_record(header, rows)
    groups = _infer_groups(header, rows, id_column, gene_column)
    annotation_bins = _infer_annotation_bins(header)
    tag_bins = _infer_tag_bins(header)
    boolean_tags = _infer_boolean_tags(header)
    presence_tags = _infer_value_presence_tags(header)

    protein_sequence_column = next((col for col in groups["protein"] if "protein_sequence" in col.lower()), "")
    mrna_sequence_column = next((col for col in groups["transcript"] if "mrna_sequence" in col.lower() or "rna_sequence" in col.lower()), "")

    schema = {
        "contract": {
            "name": CONTRACT.get("name", "functional_genomics"),
            "version": CONTRACT.get("version", 1),
            "path": CONTRACT.get("path", ""),
            "module": CONTRACT.get("module", "genomics"),
            "core_entities": list(CONTRACT.get("core_entities", [])),
            "optional_entities": list(CONTRACT.get("optional_entities", [])),
            "required_relationships": list(CONTRACT.get("required_relationships", [])),
        },
        "raw": {
            "data_path": source_file.name,
            "delimiter": "\\t" if _detect_delimiter(source_file) == "\t" else _detect_delimiter(source_file),
            "column_order": header,
        },
        "entity_model": {
            "primary_record_entity": primary_entity,
            "entities": {
                "organism": {
                    "entity_type": "organism",
                    "entity_id": f"organism:{_slug(inferred_organism)}",
                    "name": inferred_organism,
                },
                "dataset": {
                    "entity_type": "dataset",
                    "entity_id": f"dataset:{inferred_dataset_id}",
                    "name": inferred_dataset_name,
                },
                **(
                    {
                        "chromosome": {
                            "entity_type": "chromosome",
                            "source_column": "genome_location",
                            "id_template": f"chromosome:{_slug(inferred_organism)}:{{chromosome}}",
                            "name_template": "{chromosome}",
                        }
                    }
                    if "genome_location" in header
                    else {}
                ),
                "gene": {
                    "entity_type": "gene",
                    "id_column": gene_column or "gene_name",
                    "name_column": gene_column or "gene_name",
                    "metadata_columns": groups["gene"],
                },
                "transcript": {
                    "entity_type": "transcript",
                    "id_column": id_column,
                    "name_column": id_column,
                    "metadata_columns": [col for col in groups["transcript"] if col != mrna_sequence_column],
                },
                "protein": {
                    "entity_type": "protein",
                    "id_template": f"{{{id_column}}}:protein",
                    "name_template": f"{{{gene_column or id_column}}} protein",
                    "metadata_columns": [col for col in groups["protein"] if col != protein_sequence_column],
                    "sequence_column": protein_sequence_column,
                },
            },
            "relationships": list(CONTRACT.get("required_relationships", [])),
        },
        "promoted_entities": split_shared_and_specific(
            {
                **promoted_entities_for_header(dict(CONTRACT.get("promoted_entities", {})), header),
                **_infer_orthogroup_promoted_entities(header),
            },
            contract_items=CONTRACT.get("promoted_entities", {}),
        ),
        "comparative_entities": split_shared_and_specific(
            _infer_comparative_entities(header),
            contract_items=CONTRACT.get("comparative_entities", {}),
        ),
        "annotation_bins": split_shared_and_specific(annotation_bins, contract_items=CONTRACT.get("annotation_bins", [])),
        "tag_bins": split_shared_and_specific(tag_bins, contract_items=CONTRACT.get("tag_bins", [])),
        "boolean_tags": split_shared_and_specific(boolean_tags, contract_items=CONTRACT.get("boolean_tags", [])),
        "value_presence_tags": split_shared_and_specific(presence_tags, contract_items=CONTRACT.get("value_presence_tags", [])),
        "tag_hierarchy": split_shared_and_specific(dict(CONTRACT.get("tag_hierarchy", {})), contract_items=CONTRACT.get("tag_hierarchy", {})),
        "expression_entities": _infer_expression_entities(header),
    }

    if mrna_sequence_column:
        schema["entity_model"]["entities"]["transcript"]["metadata_columns"].append(mrna_sequence_column)

    dataset = {
        "dataset": {
            "id": inferred_dataset_id,
            "name": inferred_dataset_name,
            "module": "genomics",
            "extension": "functional_genomics",
            "profile": _slug(source_file.stem),
            "organism": inferred_organism,
            "description": sidecar.get("description", "Locally inferred genomics dataset metadata from a user-provided source table."),
            "raw_sources": [{"path": source_file.name, "kind": source_file.suffix.lstrip(".") or "table", "role": "primary_matrix"}],
            "standardized_sources": {"schema": "schema.yaml" if apply else "schema.inferred.yaml"},
            "primary_record": {
                "entity_type": primary_entity,
                "id_column": id_column,
                "gene_id_column": gene_column or "",
                "protein_id_template": f"{{{id_column}}}:protein",
            },
            "llm_context": {
                "summary": f"Inferred genomics dataset built from {source_file.name} with primary {primary_entity} records.",
                "glossary": {
                    "orthogroup": "Comparative gene family assignment across related organisms.",
                    "dge": "Differential gene expression represented as log2 fold-change contrasts.",
                },
            },
            "ui_hints": {
                "default_focus": "gene-centric",
                "recommended_focus_modes": [
                    "gene-centric",
                    "functional-annotation-centric",
                    "expression-centric",
                ],
            },
        }
    }

    report = {
        "source_file": str(source_file),
        "rows": len(rows),
        "columns": len(header),
        "primary_record_entity": primary_entity,
        "id_column": id_column,
        "gene_column": gene_column or "",
        "group_counts": {key: len(value) for key, value in groups.items()},
        "annotation_bins": [item["column"] for item in annotation_bins],
        "tag_bins": [item["column"] for item in tag_bins],
        "boolean_tags": [item["column"] for item in boolean_tags],
        "value_presence_tags": [item["column"] for item in presence_tags],
        "notes": [str(path) for path in notes],
    }

    dataset_name_out = "dataset.yaml" if apply else "dataset.inferred.yaml"
    schema_name_out = "schema.yaml" if apply else "schema.inferred.yaml"
    dataset_path = output_dir / dataset_name_out
    schema_path = output_dir / schema_name_out
    report_path = output_dir / "inference-report.yaml"
    dataset_path.write_text(yaml.safe_dump(dataset, sort_keys=False, allow_unicode=False))
    schema_path.write_text(yaml.safe_dump(schema, sort_keys=False, allow_unicode=False))
    report_path.write_text(yaml.safe_dump(report, sort_keys=False, allow_unicode=False))
    return dataset_path, schema_path, report_path


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Infer standardized genomics YAML sources from a raw local table.")
    parser.add_argument("--source-file", required=True)
    parser.add_argument("--source-dir", default="")
    parser.add_argument("--dataset-id", default="")
    parser.add_argument("--dataset-name", default="")
    parser.add_argument("--organism", default="")
    parser.add_argument("--notes", action="append", default=[])
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    infer_source_package(
        source_file=Path(args.source_file),
        source_dir=Path(args.source_dir) if args.source_dir else None,
        dataset_id=args.dataset_id or None,
        dataset_name=args.dataset_name or None,
        organism=args.organism or None,
        note_paths=[Path(p) for p in args.notes],
        apply=bool(args.apply),
    )
