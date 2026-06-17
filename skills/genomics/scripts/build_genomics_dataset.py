#!/usr/bin/env python3
"""Deterministic builder for genomics datasets backed by local standardized sources."""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

import yaml

from kgx.db import KnowledgeGraphDB
from genomics_contract import combine_section
from normalize_source import derive_contrast_summary_links
from render_genomics_vault import render_genomics_vault


GO_RE = re.compile(r"GO:\d+")
INTERPRO_RE = re.compile(r"IPR\d+")
PFAM_RE = re.compile(r"PF\d+")
SMART_RE = re.compile(r"SM\d+")
PANTHER_RE = re.compile(r"PTHR\d+")
FUNFAM_RE = re.compile(r"G3DSA:[^|;,\\s]+:FF:\\d+")


def _load_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text()) or {}


def _clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text or text.lower() in {"na", "n/a", "none", "null"}:
            return None
        return text
    return value


def _metadata_from_row(row: dict[str, str], columns: list[str]) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    for column in columns:
        value = _clean_value(row.get(column))
        if value is not None:
            metadata[column] = value
    return metadata


def _protein_name(row: dict[str, str], name_template: str) -> str:
    try:
        return name_template.format(**row)
    except KeyError:
        return row.get("uniquename", "protein")


def _apply_template(template: str, row: dict[str, str]) -> str:
    return template.format(**row)


def _slugify(text: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", str(text).strip().lower()).strip("-")
    return slug or "value"


def _chromosome_from_location(value: Any) -> str | None:
    cleaned = _clean_value(value)
    if cleaned is None:
        return None
    text = str(cleaned).strip()
    if not text:
        return None
    match = re.match(r"^([^:]+):", text)
    if match:
        return match.group(1).strip() or None
    match = re.search(r"(?i)\b(chr[\w.-]+|chromosome[\w .-]+|scaffold[\w.-]+|contig[\w.-]+)\b", text)
    if match:
        return match.group(1).strip()
    return None


def _expression_entity_id(spec: dict[str, Any], owner_id: str) -> str | None:
    entity_type = str(spec.get("entity_type", ""))
    id_template = str(spec.get("id_template", ""))
    if not entity_type or not id_template:
        return None
    label = str(spec.get("label") or spec.get("column") or "")
    label_slug = _slugify(label)
    return id_template.format(owner_id=owner_id, column=spec.get("column", ""), label=label, label_slug=label_slug)


def _static_entity_id(entity_cfg: dict[str, Any]) -> str | None:
    entity_id = str(entity_cfg.get("entity_id", "") or "").strip()
    return entity_id or None


def _parse_term_values(text: Any) -> list[str]:
    cleaned = _clean_value(text)
    if cleaned is None:
        return []
    return [part.strip() for part in re.split(r"[|,;]+", str(cleaned)) if part.strip()]


def _comparative_entity_id(spec: dict[str, Any], value: str) -> str:
    template = str(spec.get("id_template", "{value}"))
    normalized_value = _slugify(value) if "{value}" in template and ("comparative_hit:" in template or "homolog_gene:" in template) else value
    return template.format(value=normalized_value)


def _to_int_if_numeric(value: Any) -> Any:
    cleaned = _clean_value(value)
    if cleaned is None:
        return None
    if isinstance(cleaned, int):
        return cleaned
    text = str(cleaned).strip()
    if re.fullmatch(r"-?\d+", text):
        try:
            return int(text)
        except ValueError:
            return text
    return text


def _parse_single_term(text: str) -> list[dict[str, Any]]:
    cleaned = _clean_value(text)
    if cleaned is None:
        return []
    return [{"id": str(cleaned), "label": str(cleaned), "score": None}]


def _parse_term_list(text: str) -> list[dict[str, Any]]:
    cleaned = _clean_value(text)
    if cleaned is None:
        return []
    parts = [part.strip() for part in re.split(r"[|,;]+", str(cleaned)) if part.strip()]
    return [{"id": part, "label": part, "score": None} for part in parts]


def _parse_go_plain(text: str) -> list[dict[str, Any]]:
    cleaned = _clean_value(text)
    if cleaned is None:
        return []
    return [{"id": match, "label": match, "score": None} for match in sorted(set(GO_RE.findall(str(cleaned))))]


def _parse_go_scored(text: str) -> list[dict[str, Any]]:
    cleaned = _clean_value(text)
    if cleaned is None:
        return []
    results: list[dict[str, Any]] = []
    for part in str(cleaned).split("|"):
        token = part.strip()
        if not token or "_" not in token:
            continue
        go_id, raw_score = token.rsplit("_", 1)
        if not GO_RE.fullmatch(go_id):
            continue
        try:
            score = float(raw_score)
        except ValueError:
            score = None
        results.append({"id": go_id, "label": go_id, "score": score})
    dedup: dict[str, dict[str, Any]] = {}
    for item in results:
        prev = dedup.get(item["id"])
        if prev is None or (item["score"] or 0.0) > (prev["score"] or 0.0):
            dedup[item["id"]] = item
    return list(dedup.values())


def _regex_parser(pattern: re.Pattern[str], text: str) -> list[dict[str, Any]]:
    cleaned = _clean_value(text)
    if cleaned is None:
        return []
    return [{"id": match, "label": match, "score": None} for match in sorted(set(pattern.findall(str(cleaned))))]


PARSERS = {
    "single_term": _parse_single_term,
    "term_list": _parse_term_list,
    "go_plain": _parse_go_plain,
    "go_scored": _parse_go_scored,
    "interpro_ids": lambda text: _regex_parser(INTERPRO_RE, text),
    "pfam_ids": lambda text: _regex_parser(PFAM_RE, text),
    "smart_ids": lambda text: _regex_parser(SMART_RE, text),
    "panther_ids": lambda text: _regex_parser(PANTHER_RE, text),
    "funfam_ids": lambda text: _regex_parser(FUNFAM_RE, text),
}


def _slug_label(raw_id: str) -> str:
    label = str(raw_id).replace("_", " ").replace("-", " ").strip()
    return " ".join(part.capitalize() if part.islower() else part for part in label.split())


def _ensure_tag(db: KnowledgeGraphDB, tag_id: str, *, name: str, category: str, parent: str = "", metadata: dict[str, Any] | None = None) -> str:
    payload = {"category": category}
    if metadata:
        payload.update(metadata)
    canonical = db.upsert_entity("tag", tag_id, name=name, metadata=payload)
    if parent:
        db.add_relationship(canonical, "BROADER", parent)
    return canonical


def _seed_tag_hierarchy(db: KnowledgeGraphDB, hierarchy: dict[str, dict[str, Any]]) -> dict[str, str]:
    created: dict[str, str] = {}
    for tag_id, info in hierarchy.items():
        parent = info.get("parent", "") or ""
        created[tag_id] = _ensure_tag(
            db,
            tag_id,
            name=info.get("name", _slug_label(tag_id)),
            category=info.get("category", "topic"),
            parent=created.get(parent, parent),
        )
    return created


def _attach_tag(
    db: KnowledgeGraphDB,
    *,
    entity_id: str,
    item: dict[str, Any],
    parent_tag: str,
    namespace: str,
    relation_metadata: dict[str, Any] | None = None,
) -> None:
    raw_id = item["id"]
    if namespace and not str(raw_id).lower().startswith(f"{namespace.lower()}:"):
        entity_key = f"{namespace}:{raw_id}"
    else:
        entity_key = str(raw_id)
    tag_name = item.get("label") or str(raw_id)
    tag_metadata = {
        "category": "topic",
        "namespace": namespace,
    }
    canonical = _ensure_tag(db, entity_key, name=tag_name, category="topic", parent=parent_tag, metadata=tag_metadata)
    payload = dict(relation_metadata or {})
    if item.get("score") is not None:
        payload["score"] = item["score"]
    db.add_relationship(entity_id, "TAGGED", canonical, metadata=payload)


def _promote_linked_entity(
    db: KnowledgeGraphDB,
    *,
    source_entity_id: str,
    source_entity_type: str,
    item: dict[str, Any],
    spec: dict[str, Any],
) -> str | None:
    entity_type = spec.get("promoted_entity_type", "")
    relation_type = spec.get("promoted_relation_type", "")
    id_template = spec.get("promoted_id_template", "")
    if not entity_type or not relation_type or not id_template:
        return None

    namespace = str(spec.get("namespace", ""))
    raw_id = str(item.get("id", ""))
    bare_id = raw_id
    if namespace and raw_id.lower().startswith(f"{namespace.lower()}:"):
        bare_id = raw_id[len(namespace) + 1 :]

    values = {
        "id": bare_id,
        "raw_id": raw_id,
        "label": str(item.get("label") or item.get("id") or ""),
        "score": item.get("score"),
        "namespace": namespace,
        "column": str(spec.get("column", "")),
        "category": str(spec.get("category", "")),
        "attach_to": str(spec.get("attach_to", source_entity_type)),
    }
    entity_id = id_template.format(**values)
    metadata = {
        "category": values["category"],
        "source_column": values["column"],
        "source_entity_type": source_entity_type,
    }
    if values["namespace"]:
        metadata["namespace"] = values["namespace"]
    if values["score"] is not None:
        metadata["score"] = values["score"]

    promoted_id = db.upsert_entity(
        entity_type,
        entity_id,
        name=values["label"] or values["id"] or entity_id,
        metadata=metadata,
    )
    db.add_relationship(source_entity_id, relation_type, promoted_id)
    parent_tag = str(spec.get("parent_tag", "") or "")
    if parent_tag:
        db.add_relationship(promoted_id, "TAGGED", parent_tag, metadata={"source_column": values["column"]})
    return promoted_id


def _promote_expression_entity(
    db: KnowledgeGraphDB,
    *,
    source_entity_id: str,
    owner_id: str,
    row: dict[str, Any],
    spec: dict[str, Any],
) -> str | None:
    column = str(spec.get("column", ""))
    value = _clean_value(row.get(column))
    if not column or value is None:
        return None
    entity_type = str(spec.get("entity_type", ""))
    relation_type = str(spec.get("relation_type", ""))
    id_template = str(spec.get("id_template", ""))
    if not entity_type or not relation_type or not id_template:
        return None
    label = str(spec.get("label") or column)
    value_key = str(spec.get("value_key", "value"))
    label_slug = _slugify(label)
    entity_id = id_template.format(owner_id=owner_id, column=column, label=label, label_slug=label_slug)
    metadata = {
        "category": str(spec.get("measure_type", "")),
        "source_column": column,
        "label": label,
    }
    order_index = spec.get("order_index")
    if order_index is not None:
        metadata["order_index"] = order_index
        if metadata["category"] == "summary":
            metadata["stage_order"] = order_index
        elif metadata["category"] == "contrast":
            metadata["contrast_order"] = order_index
    for key in ("source_summary_label", "source_summary_column", "target_summary_label", "target_summary_column"):
        if spec.get(key):
            metadata[key] = spec[key]
    promoted_id = db.upsert_entity(
        entity_type,
        entity_id,
        name=label,
        metadata=metadata,
    )
    db.add_relationship(source_entity_id, relation_type, promoted_id, metadata={value_key: value, "source_column": column})
    parent_tag = str(spec.get("parent_tag", "") or "")
    if parent_tag:
        db.add_relationship(promoted_id, "TAGGED", parent_tag, metadata={"source_column": column})
    return promoted_id


def _iter_rows(data_path: Path, delimiter: str) -> list[dict[str, Any]]:
    suffix = data_path.suffix.lower()
    if suffix == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(data_path, read_only=True, data_only=True)
        ws = wb.active
        values = list(ws.iter_rows(values_only=True))
        if not values:
            return []
        header = [str(cell).strip() if cell is not None else "" for cell in values[0]]
        rows: list[dict[str, Any]] = []
        for row_values in values[1:]:
            row = {}
            for idx, col in enumerate(header):
                if not col:
                    continue
                row[col] = row_values[idx] if idx < len(row_values) else None
            rows.append(row)
        return rows

    actual_delimiter = "\t" if delimiter == "\\t" else delimiter
    with data_path.open(newline="") as handle:
        return list(csv.DictReader(handle, delimiter=actual_delimiter))


def build_dataset(*, source_dir: Path, db_path: Path, fresh: bool = False, vault_output_dir: Path | None = None) -> Path:
    source_dir = source_dir.resolve()
    dataset_cfg = _load_yaml(source_dir / "dataset.yaml").get("dataset", {})
    schema = _load_yaml(source_dir / "schema.yaml")
    raw_cfg = schema.get("raw", {})
    entity_model = schema.get("entity_model", {})
    entities_cfg = entity_model.get("entities", {})

    data_path = source_dir / raw_cfg.get("data_path", "DATA.tsv")
    rows = _iter_rows(data_path, raw_cfg.get("delimiter", "\\t"))

    if fresh and db_path.exists():
        db_path.unlink()
    db_path.parent.mkdir(parents=True, exist_ok=True)

    organism_entity_cfg = entities_cfg["organism"]
    dataset_entity_cfg = entities_cfg["dataset"]
    annotation_bins = combine_section(schema.get("annotation_bins", []))
    tag_bins = combine_section(schema.get("tag_bins", []))
    boolean_tags = combine_section(schema.get("boolean_tags", []))
    value_presence_tags = combine_section(schema.get("value_presence_tags", []))
    comparative_entities = combine_section(schema.get("comparative_entities", {}))
    expression_entities = schema.get("expression_entities", {}) or {}
    expression_summaries = combine_section(expression_entities.get("summaries", []))
    expression_contrasts = combine_section(expression_entities.get("contrasts", []))
    ui_cfg = schema.get("ui", {}) or {}
    expression_field_lookup = {
        str(label): str(column)
        for label, column in (ui_cfg.get("expression_fields", {}) or {}).items()
    }
    contrast_field_lookup = {
        str(label): str(column)
        for label, column in (ui_cfg.get("log2fc_fields", {}) or {}).items()
    }
    summary_spec_by_column = {
        str(spec.get("column", "")): spec
        for spec in expression_summaries
        if spec.get("column")
    }
    orthogroup_cfg = combine_section(schema.get("promoted_entities", {})).get("orthogroup", {})
    orthogroup_state: dict[str, dict[str, Any]] = {}
    comparative_state: dict[str, dict[str, Any]] = {}

    with KnowledgeGraphDB(db_path) as db:
        seeded_tags = _seed_tag_hierarchy(db, combine_section(schema.get("tag_hierarchy", {})))
        linked_contrast_edges: set[tuple[str, str, str]] = set()
        comparative_organism_ids: dict[str, str] = {}
        comparative_identity_index: dict[tuple[str, str], str] = {}
        organism_entity_id = db.upsert_entity(
            organism_entity_cfg.get("entity_type", "organism"),
            _static_entity_id(organism_entity_cfg) or f"organism:{_slugify(dataset_cfg.get('organism', 'unknown-organism'))}",
            name=organism_entity_cfg.get("name", dataset_cfg.get("organism", "Unknown organism")),
            metadata={
                "label": dataset_cfg.get("organism", ""),
                "datasets": [dataset_cfg.get("id", "")] if dataset_cfg.get("id") else [],
            },
        )
        comparative_organism_ids[str(dataset_cfg.get("organism", "") or organism_entity_cfg.get("name", ""))] = organism_entity_id

        dataset_entity_id = db.upsert_entity(
            dataset_entity_cfg.get("entity_type", "dataset"),
            dataset_entity_cfg["entity_id"],
            name=dataset_entity_cfg["name"],
            metadata={
                "module": dataset_cfg.get("module", "genomics"),
                "extension": dataset_cfg.get("extension", "functional_genomics"),
                "profile": dataset_cfg.get("profile", ""),
                "organism": dataset_cfg.get("organism", ""),
                "description": dataset_cfg.get("description", ""),
            },
        )
        db.add_relationship(dataset_entity_id, "ABOUT_ORGANISM", organism_entity_id)
        chromosome_cfg = entities_cfg.get("chromosome", {})

        for row in rows:
            gene_cfg = entities_cfg["gene"]
            transcript_cfg = entities_cfg["transcript"]
            protein_cfg = entities_cfg["protein"]

            gene_name = _clean_value(row.get(gene_cfg["id_column"]))
            transcript_name = _clean_value(row.get(transcript_cfg["id_column"]))
            if gene_name is None or transcript_name is None:
                continue

            gene_id = db.upsert_entity(
                gene_cfg["entity_type"],
                str(gene_name),
                name=str(gene_name),
                metadata=_metadata_from_row(row, gene_cfg.get("metadata_columns", [])),
            )
            transcript_id = db.upsert_entity(
                transcript_cfg["entity_type"],
                str(transcript_name),
                name=str(transcript_name),
                metadata=_metadata_from_row(row, transcript_cfg.get("metadata_columns", [])),
            )

            protein_id_value = _apply_template(protein_cfg["id_template"], row)
            protein_metadata = _metadata_from_row(row, protein_cfg.get("metadata_columns", []))
            sequence_column = protein_cfg.get("sequence_column", "")
            sequence = _clean_value(row.get(sequence_column))
            if sequence is not None:
                protein_metadata["protein_sequence"] = sequence
                protein_metadata["length"] = len(str(sequence))
            protein_id = db.upsert_entity(
                protein_cfg["entity_type"],
                protein_id_value,
                name=_protein_name(row, protein_cfg["name_template"]),
                metadata=protein_metadata,
            )

            db.add_relationship(gene_id, "HAS_TRANSCRIPT", transcript_id)
            db.add_relationship(transcript_id, "TRANSLATED_TO", protein_id)
            db.add_relationship(gene_id, "IN_DATASET", dataset_entity_id)
            db.add_relationship(gene_id, "FROM_ORGANISM", organism_entity_id)

            chromosome_entity_id = None
            chromosome_source_column = str(chromosome_cfg.get("source_column", "") or "")
            chromosome_label = _chromosome_from_location(row.get(chromosome_source_column)) if chromosome_source_column else None
            if chromosome_label is not None:
                chromosome_entity_id = db.upsert_entity(
                    chromosome_cfg.get("entity_type", "chromosome"),
                    chromosome_cfg.get(
                        "id_template",
                        f"chromosome:{_slugify(dataset_cfg.get('organism', 'unknown-organism'))}:{{chromosome}}",
                    ).format(chromosome=_slugify(chromosome_label)),
                    name=chromosome_cfg.get("name_template", "{chromosome}").format(chromosome=chromosome_label),
                    metadata={
                        "label": chromosome_label,
                        "organism": dataset_cfg.get("organism", ""),
                        "source_column": chromosome_source_column,
                    },
                )
                db.add_relationship(organism_entity_id, "HAS_CHROMOSOME", chromosome_entity_id)
                db.add_relationship(chromosome_entity_id, "HAS_GENE", gene_id)

            orthogroup_id = None
            orthogroup_value = _clean_value(row.get(orthogroup_cfg.get("source_column", "")))
            if orthogroup_value is not None:
                orthogroup_metadata = _metadata_from_row(row, orthogroup_cfg.get("metadata_columns", []))
                orthogroup_id = db.upsert_entity(
                    orthogroup_cfg.get("entity_type", "orthogroup"),
                    orthogroup_cfg.get("id_template", "orthogroup:{value}").format(value=orthogroup_value),
                    name=orthogroup_cfg.get("name_template", "{value}").format(value=orthogroup_value),
                    metadata={**orthogroup_metadata, "dataset": dataset_cfg.get("id", "")},
                )
                db.add_relationship(gene_id, orthogroup_cfg.get("relationship_type", "BELONGS_TO_ORTHOGROUP"), orthogroup_id)
                state = orthogroup_state.setdefault(orthogroup_id, {
                    "name": orthogroup_cfg.get("name_template", "{value}").format(value=orthogroup_value),
                    "row_metadata": {},
                    "local_gene_ids": set(),
                    "schachtii_genes": set(),
                })
                for key, value in orthogroup_metadata.items():
                    if key == "schachtii_genes":
                        state["schachtii_genes"].update(_parse_term_values(value))
                    else:
                        normalized_value = _to_int_if_numeric(value)
                        if normalized_value is not None:
                            state["row_metadata"][key] = normalized_value
                state["local_gene_ids"].add(gene_id)

            entity_ids = {
                "organism": organism_entity_id,
                "dataset": dataset_entity_id,
                "chromosome": chromosome_entity_id,
                "gene": gene_id,
                "transcript": transcript_id,
                "protein": protein_id,
                "orthogroup": orthogroup_id,
            }

            for spec in comparative_entities.values():
                source_value = row.get(spec.get("source_column", ""))
                if not source_value:
                    continue
                attach_from = str(spec.get("attach_from", "") or "")
                attach_id = entity_ids.get(attach_from)
                if not attach_id:
                    continue
                for item_value in _parse_term_values(source_value):
                    target_organism = str(spec.get("target_organism", "") or "").strip()
                    identity_key = (target_organism.lower(), item_value.strip().lower()) if target_organism else ("", item_value.strip().lower())
                    comparative_id = ""
                    if bool(spec.get("reuse_target_organism_identity")) and identity_key in comparative_identity_index:
                        comparative_id = comparative_identity_index[identity_key]
                    if not comparative_id:
                        comparative_id = _comparative_entity_id(spec, item_value)

                    comparative_organism_id = ""
                    if target_organism and str(spec.get("entity_type", "")) == "bcn_gene":
                        comparative_organism_id = comparative_organism_ids.get(target_organism, "")
                        if not comparative_organism_id:
                            comparative_organism_id = db.upsert_entity(
                                "organism",
                                f"organism:{_slugify(target_organism)}",
                                name=target_organism,
                                metadata={"label": target_organism, "datasets": []},
                            )
                            comparative_organism_ids[target_organism] = comparative_organism_id
                    db.upsert_entity(
                        str(spec.get("entity_type", "comparative_hit") or "comparative_hit"),
                        comparative_id,
                        name=str(spec.get("name_template", "{value}")).format(value=item_value),
                        metadata={"category": "comparative"},
                    )
                    if target_organism and identity_key[1]:
                        comparative_identity_index.setdefault(identity_key, comparative_id)
                    state = comparative_state.setdefault(comparative_id, {
                        "name": str(spec.get("name_template", "{value}")).format(value=item_value),
                        "entity_type": str(spec.get("entity_type", "comparative_hit") or "comparative_hit"),
                        "organism": str(spec.get("target_organism", "") or ""),
                        "source_columns": set(),
                        "relationship_types": set(),
                        "scope_tag_ids": set(),
                    })
                    state["source_columns"].add(str(spec.get("source_column", "") or ""))
                    state["relationship_types"].add(str(spec.get("relationship_type", "") or ""))
                    db.add_relationship(
                        attach_id,
                        str(spec.get("relationship_type", "RELATED") or "RELATED"),
                        comparative_id,
                        metadata={"source_column": str(spec.get("source_column", "") or "")},
                    )
                    if comparative_organism_id:
                        db.add_relationship(comparative_id, "FROM_ORGANISM", comparative_organism_id)
                    scope_tag_id = str(spec.get("scope_tag_id", "") or "").strip()
                    if scope_tag_id:
                        canonical_scope_tag = seeded_tags.get(scope_tag_id, scope_tag_id)
                        db.add_relationship(
                            comparative_id,
                            "TAGGED",
                            canonical_scope_tag,
                            metadata={"source_column": str(spec.get("source_column", "") or "")},
                        )
                        state["scope_tag_ids"].add(canonical_scope_tag)

            for spec in annotation_bins:
                parser = PARSERS[spec["parser"]]
                items = parser(row.get(spec["column"], ""))
                if not items:
                    continue
                attach_id = entity_ids[spec["attach_to"]]
                parent_tag = seeded_tags.get(spec["parent_tag"], spec["parent_tag"])
                for item in items:
                    promoted_id = _promote_linked_entity(
                        db,
                        source_entity_id=attach_id,
                        source_entity_type=spec["attach_to"],
                        item=item,
                        spec={**spec, "parent_tag": parent_tag},
                    )
                    if promoted_id is None:
                        _attach_tag(
                            db,
                            entity_id=attach_id,
                            item=item,
                            parent_tag=parent_tag,
                            namespace=spec.get("namespace", ""),
                            relation_metadata={"source_column": spec["column"]},
                        )

            for spec in tag_bins:
                parser = PARSERS[spec["parser"]]
                items = parser(row.get(spec["column"], ""))
                if not items:
                    continue
                attach_id = entity_ids[spec["attach_to"]]
                parent_tag = seeded_tags.get(spec["parent_tag"], spec["parent_tag"])
                for item in items:
                    normalized = {
                        "id": str(item["id"]).strip().lower().replace(" ", "_"),
                        "label": str(item["label"]).strip(),
                        "score": item.get("score"),
                    }
                    promoted_id = _promote_linked_entity(
                        db,
                        source_entity_id=attach_id,
                        source_entity_type=spec["attach_to"],
                        item=normalized,
                        spec={**spec, "parent_tag": parent_tag},
                    )
                    if promoted_id is None:
                        _attach_tag(
                            db,
                            entity_id=attach_id,
                            item=normalized,
                            parent_tag=parent_tag,
                            namespace="tag",
                            relation_metadata={"source_column": spec["column"]},
                        )

            for spec in boolean_tags:
                cleaned = str(_clean_value(row.get(spec["column"])) or "").lower()
                if cleaned not in set(spec.get("truthy", [])):
                    continue
                parent_tag = seeded_tags.get(spec["parent_tag"], spec["parent_tag"])
                _attach_tag(
                    db,
                    entity_id=entity_ids[spec["attach_to"]],
                    item={"id": spec["tag_id"], "label": spec["label"], "score": None},
                    parent_tag=parent_tag,
                    namespace="tag",
                    relation_metadata={"source_column": spec["column"]},
                )

            for spec in value_presence_tags:
                if _clean_value(row.get(spec["column"])) is None:
                    continue
                parent_tag = seeded_tags.get(spec["parent_tag"], spec["parent_tag"])
                _attach_tag(
                    db,
                    entity_id=entity_ids[spec["attach_to"]],
                    item={"id": spec["tag_id"], "label": spec["label"], "score": None},
                    parent_tag=parent_tag,
                    namespace="tag",
                    relation_metadata={"source_column": spec["column"]},
                )

            for spec in expression_summaries:
                attach_to = str(spec.get("attach_to", "transcript"))
                owner_entity_id = entity_ids.get(attach_to)
                if owner_entity_id:
                    _promote_expression_entity(
                        db,
                        source_entity_id=owner_entity_id,
                        owner_id=owner_entity_id,
                        row=row,
                        spec=spec,
                    )

            for spec in expression_contrasts:
                attach_to = str(spec.get("attach_to", "transcript"))
                owner_entity_id = entity_ids.get(attach_to)
                if owner_entity_id:
                    contrast_entity_id = _promote_expression_entity(
                        db,
                        source_entity_id=owner_entity_id,
                        owner_id=owner_entity_id,
                        row=row,
                        spec=spec,
                    )
                    if not contrast_entity_id:
                        continue
                    contrast_spec = dict(spec)
                    if not contrast_spec.get("source_summary_column") and expression_field_lookup:
                        contrast_spec.update(
                            derive_contrast_summary_links(
                                expression_field_lookup,
                                str(spec.get("label") or ""),
                                contrast_field_lookup.get(str(spec.get("label") or ""), str(spec.get("column") or "")),
                            )
                        )
                    for rel_type, column_key, label_key in (
                        ("CONTRAST_SOURCE", "source_summary_column", "source_summary_label"),
                        ("CONTRAST_TARGET", "target_summary_column", "target_summary_label"),
                    ):
                        summary_column = str(contrast_spec.get(column_key) or "").strip()
                        if not summary_column:
                            continue
                        summary_spec = summary_spec_by_column.get(summary_column)
                        if not summary_spec:
                            continue
                        summary_entity_id = _expression_entity_id(summary_spec, owner_entity_id)
                        if not summary_entity_id:
                            continue
                        edge_key = (contrast_entity_id, rel_type, summary_entity_id)
                        if edge_key in linked_contrast_edges:
                            continue
                        db.add_relationship(
                            contrast_entity_id,
                            rel_type,
                            summary_entity_id,
                            metadata={
                                "contrast_column": str(spec.get("column") or ""),
                                "summary_column": summary_column,
                                "summary_label": str(contrast_spec.get(label_key) or ""),
                            },
                        )
                        linked_contrast_edges.add(edge_key)

        for orthogroup_id, state in orthogroup_state.items():
            member_gene_ids = sorted(state["local_gene_ids"])
            metadata = {
                **state["row_metadata"],
                "dataset": dataset_cfg.get("id", ""),
                "local_gene_count": len(member_gene_ids),
                "local_gene_ids": member_gene_ids,
            }
            if state["schachtii_genes"]:
                metadata["schachtii_genes"] = sorted(state["schachtii_genes"])
            db.upsert_entity(
                orthogroup_cfg.get("entity_type", "orthogroup"),
                orthogroup_id,
                name=state["name"],
                metadata=metadata,
            )

        for comparative_id, state in comparative_state.items():
            metadata = {
                "category": "comparative",
                "organism": state["organism"],
                "source_columns": sorted(column for column in state["source_columns"] if column),
                "relationship_types": sorted(rel for rel in state["relationship_types"] if rel),
            }
            if state["scope_tag_ids"]:
                metadata["scope_tag_ids"] = sorted(state["scope_tag_ids"])
            if len(metadata["source_columns"]) == 1:
                metadata["source_column"] = metadata["source_columns"][0]
            db.upsert_entity(
                state["entity_type"],
                comparative_id,
                name=state["name"],
                metadata=metadata,
            )

    if vault_output_dir is not None:
        render_genomics_vault(db_path=db_path, output_dir=vault_output_dir)

    return db_path


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Build a genomics SQLite DB from standardized YAML sources.")
    parser.add_argument("--source-dir", required=True)
    parser.add_argument("--db", required=True)
    parser.add_argument("--fresh", action="store_true")
    parser.add_argument("--vault-output", default="")
    cli_args = parser.parse_args()
    build_dataset(
        source_dir=Path(cli_args.source_dir),
        db_path=Path(cli_args.db),
        fresh=cli_args.fresh,
        vault_output_dir=Path(cli_args.vault_output) if cli_args.vault_output else None,
    )
